import torch
from torch import nn, save
from torch.optim import Adam
import torch.utils.data as Data
from torchvision import  transforms
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from vit import VisionTransformer
import new_loader

root = 'Malnet/'

pathTrain = 'train/'
pathValid = 'val/'


def datasaver(v_accu_e, t_accu_e, clf):
        #Saving trained model
    with open('model.cnn', 'wb') as f:
        save(clf.state_dict(), f)

    #Saving accuracy info
    wb = openpyxl.Workbook()
    sheet = wb.active

    aCell = sheet.cell(row= 1, column=1)
    aCell.value = "Generation"
    aCell = sheet.cell(row= 1, column=2)
    aCell.value = "Training"
    aCell = sheet.cell(row= 1, column=3)
    aCell.value = "Validation"

    for i in range(len(v_accu_e)):
        aCell = sheet.cell(row= i+2, column=1)
        aCell.value = i+1
        aCell = sheet.cell(row= i+2, column=2)
        aCell.value = t_accu_e[i]
        aCell = sheet.cell(row= i+2, column=3)
        aCell.value = v_accu_e[i]

    wb.save("accuracy.xlsx")
    

    #Plotting plots for accuracies over epochs
    plt.title("Accuracy vs Epochs")
    plt.xlabel("epoch")
    plt.ylabel("accuracy")
    plt.plot(v_accu_e, label = "validation" , color = 'red')
    plt.plot(t_accu_e, label = "training" , color = 'blue')
    plt.axvline(x = np.argmax(v_accu_e), label = "highest validation accuracy", color = 'red',linestyle='dashed')
    plt.axvline(x = np.argmax(t_accu_e), label = "highest training accuracy", color = 'blue',linestyle='dashed')
    leg = plt.legend(loc='center right')
    plt.savefig("graph.png")

#parameters
bsize = 1
EP = 10
lrate = 0.00001
limit = 0.0000001

transform = transforms.Compose([
    transforms.Resize((256, 256)),  # Resize the image to a specific size
    transforms.ToTensor(),  # Convert image to tensor
    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])  # Normalize image
])

train = new_loader.ImageFolderDataset(root+pathTrain, transform=transform)
val = new_loader.ImageFolderDataset(root+pathValid, transform=transform)

trainset = Data.DataLoader (dataset=train, batch_size=bsize, shuffle=True)
valset = Data.DataLoader (dataset=val, batch_size=bsize, shuffle=True)

clf = VisionTransformer().to('cuda')  
opt = Adam(clf.parameters(), lr=lrate)
lrsc = torch.optim.lr_scheduler.ReduceLROnPlateau(optimizer=opt, mode='max', cooldown=int(EP/10), verbose=True, min_lr=limit)
loss_fn = nn.CrossEntropyLoss()

v_accu_e = []
t_accu_e = []

if __name__ == "__main__":

    for epoch in range(EP): #specifying number of epochs
        print(f"Epoch:[{epoch}]")

        ######################################################################
        accu = 0

        
        for s, batch in enumerate(valset): #interating for all batches in validation set
            X,y = batch
            X, y = X.to('cuda'), y.to('cuda')
            yhat = clf(X)

            #Finding validation accuracy for epoch
            for i, pred in enumerate(yhat.cpu().detach().numpy()):
                if(np.argmax(pred) == y.cpu().detach().numpy()[i]):
                    accu = accu + 1

        #Validation accuracy, loss over epochs
        accu = accu / (len(valset) * bsize)

        v_accu_e.append(accu)        
        

        ######################################################################
        accu = 0

        for s, batch in enumerate(trainset): #interating for all batches in training set
            
            X,y = batch
            X, y = X.to('cuda'), y.to('cuda')
            yhat = clf(X)
            #print(yhat.shape)
            Y = torch.tensor([[0]*43],dtype=torch.float32)
            Y[0,y] = 1
            Y = Y.to('cuda')

            loss = loss_fn(yhat, Y)

            #Finding training accuracy for epoch
            for i, pred in enumerate(yhat.cpu().detach().numpy()):
                if(np.argmax(pred) == y.cpu().detach().numpy()[i]):
                    accu = accu + 1

            #Backprop + Optimization
            opt.zero_grad()
            loss.backward()
            opt.step()

        #Training accuracy over epochs
        accu = accu / (len(trainset) * bsize)

        t_accu_e.append(accu)

        lrsc.step(v_accu_e[i])
        
        ######################################################################
        
    datasaver(v_accu_e, t_accu_e, clf)
    ##########################################################################